#!/usr/bin/python

import vobject
from openpyxl import load_workbook
import re
import os
from datetime import datetime


def isNotBlank(string):
    return bool(string) and bool(string.strip())

excelfile = 'Kontakte.xlsx'

# update vcards only if the excel file is newer
last_excel_modification = datetime.fromtimestamp(os.path.getmtime(excelfile))
current_time = datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')

wb = load_workbook(filename=excelfile, read_only=True)
sheet = wb.active

# save the last one, so combined records could be build
last_contact = None

for row in sheet.iter_rows(min_row=2):
    # assume that the row could be empty
    valid_row = False

    # check all cells of the row if a single value exists
    for cell in row:
        if isinstance(cell.value, datetime) or isinstance(cell.value, long) or isNotBlank(cell.value):
            valid_row = True

    # only Rows with a single value are interesting
    if valid_row:
        # copy raw values from excel row
        full_name = row[0].value
        full_address = row[1].value
        mobile_phone = row[2].value
        home_phone = row[3].value
        work_phone = row[4].value
        email = row[5].value
        birthday = row[6].value
        group_name = row[7].value
        notes = row[8].value

        # init contact details
        given_name = None
        family_name = ''
        city_name = ''
        street_name = ''
        zip_code = ''

        # split complex values
        if isNotBlank(full_name):
            # Create new contact
            new_contact = vobject.vCard()

            # "name" or a name with the pattern "lastname, firstname"
            if ',' in full_name:
                parts = full_name.split(',')
                family_name = parts[0].strip()
                given_name = parts[1].strip()
                new_contact.add('fn').value = given_name + " " + family_name
            else:
                given_name = full_name
                new_contact.add('fn').value = full_name
        else:
            # append on the last contact
            new_contact = last_contact
            full_name = last_contact.fn.value

        file_name = re.sub("[^a-zA-Z]", "", "".join(new_contact.fn.value.split())) + '.vcf'

        # fix phone numbers
        if isNotBlank(mobile_phone):
            mobile_phone = re.sub("[^0-9]", "", mobile_phone)
            if mobile_phone.startswith('00'):
                mobile_phone = '+' + mobile_phone[2:]
            if mobile_phone.startswith('0'):
                mobile_phone = '+49' + mobile_phone[1:]
        if isNotBlank(home_phone):
            home_phone = re.sub("[^0-9]", "", home_phone)
            if home_phone.startswith('00'):
                home_phone = '+' + home_phone[2:]
            if home_phone.startswith('0'):
                home_phone = '+49' + home_phone[1:]
        if isNotBlank(work_phone):
            work_phone = re.sub("[^0-9]", "", work_phone)
            if work_phone.startswith('00'):
                work_phone = '+' + work_phone[2:]
            if work_phone.startswith('0'):
                work_phone = '+49' + work_phone[1:]

        # fill contact
        if isNotBlank(given_name):
            new_contact.add('n').value = vobject.vcard.Name(family=family_name, given=given_name)

        if isNotBlank(email):
            new_contact.add('email').value = email
            new_contact.email.type_param = 'INTERNET'

        if isNotBlank(full_address):
            # split the address as "street, city" or "city"
            if ',' in full_address:
                parts = full_address.split(',')
                street_name = parts[0].strip()
                city_name = parts[1].strip()

                # if a space in the city name, its assume that the first part is a zipcode
                if ' ' in city_name:
                    parts = city_name.split(' ')
                    zip_code = parts[0].strip()
                    city_name = parts[1].strip()
            else:
                city_name = full_address

            new_contact.add('adr').value = vobject.vcard.Address(street=street_name, city=city_name, code=zip_code)

        if isNotBlank(mobile_phone):
            mobile = new_contact.add('tel')
            mobile.value = mobile_phone
            mobile.type_param = 'CELL'

        if isNotBlank(home_phone):
            home = new_contact.add('tel')
            home.value = home_phone
            home.type_param = 'HOME'

        if isNotBlank(work_phone):
            work = new_contact.add('tel')
            work.value = work_phone
            work.type_param = 'WORK'

        if isinstance(birthday, datetime):
            new_contact.add('bday').value = birthday.strftime('%Y-%m-%d')
        elif isinstance(birthday, str):
            new_contact.add('bday').value = datetime.strptime(birthday, '%d.%m.').strftime('--%m-%d')

        # optional grouping
        if isNotBlank(group_name):
            new_contact.add('categories').value = [group_name]

        # custom notes
        if isNotBlank(notes):
            new_contact.add('note').value = notes

        if not hasattr(new_contact, 'ref'):
            new_contact.add('ref').value = current_time

        # write the card
        grouping_directory = 'contacts'
        if not os.path.exists(grouping_directory):
            os.makedirs(grouping_directory)
        f = open(grouping_directory + '/' + file_name, 'w')
        f.write(new_contact.serialize())
        f.close()

        last_contact = new_contact
